import os
import uuid
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from api.assessment.constants import ROLES, COLOR_MAPPING

from api.assessment.utils.common import generate_user_pie_chart,generate_jobs_pie_chart
import matplotlib
matplotlib.use('agg')
from matplotlib import pyplot as plt
from matplotlib.patheffects import withStroke
import matplotlib.patheffects as path_effects
import numpy as np
import excel2img
import img2pdf
from api.constants import API_DIR


activities_path = os.path.join(API_DIR, "assessment","assets","level2", 'dimmensions')
assets_folder = os.path.join(API_DIR, "assessment","assets","level1")

emotions_path = os.path.join(API_DIR, "assessment","assets","level2", 'emotions')
virtues_path = os.path.join(API_DIR, "assessment","assets","level2", 'virtues')
graphics_path = os.path.join(API_DIR, "assessment","assets","level2", 'graphics')


def find_index_of_dimmension(data, target_name):
    for i, sublist in enumerate(data):
        for item in sublist:
            if item['name'] == target_name:
                if i == 0:
                    return "Most Preferred"
                elif i == 1:
                    return "Lesser Preferred"
                elif i == 2:
                    return "Least Preferred"
                else:
                    return f"Preference Level {i}"
    return None

def add_image_to_worksheet(worksheet, folder_path, image_filename, row, column, width, height):
    img = Image(os.path.join(folder_path, image_filename))
    cell = worksheet.cell(row=row, column=column)
    img.width = width
    img.height = height
    worksheet.add_image(img, cell.coordinate)

def create_output_folder(folder_name):
    folder_path = os.path.join(API_DIR, "assessment","reports","level2", folder_name)
    os.makedirs(folder_path)
    return folder_path

def update_worksheet_cells(worksheet, replacements):
    for cell_reference, replacement_value in replacements.items():
        
        if isinstance(replacement_value, list):
            param,newValue = replacement_value
            worksheet[cell_reference] = worksheet[cell_reference].value.replace(param,newValue)

        else :
            worksheet[cell_reference] = replacement_value



def Generate_level2_Report(res,nlp_data,bucket_mapping,user_profile):
    
    random_hash = str(uuid.uuid4().hex)
    new_folder_path = create_output_folder(random_hash)
    
    excel = insert_image_into_excel(worksheet_name='Page2',data=res)
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page3',data=[res[2],nlp_data])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page4',data=[res[0][0],bucket_mapping,'power'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page5',data=[res[0][1],bucket_mapping,'power'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page6',data=[res[0][2],bucket_mapping,'power'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page7',data=[res[1][0],bucket_mapping,'push'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page8',data=[res[1][1],bucket_mapping,'push'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page9',data=[res[1][2],bucket_mapping,'push'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page10',data=[res[2][0],bucket_mapping,'pain'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page11',data=[res[2][1],bucket_mapping,'pain'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page12',data=[res[2][2],bucket_mapping,'pain'])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page13',data=[new_folder_path,user_profile,res])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page14',data=[new_folder_path,user_profile,res])
    excel.save(os.path.join(new_folder_path, "level2report.xlsx"))

    convert_excel_to_pdf(new_folder_path,"type")

    return os.path.join(new_folder_path, "output.pdf")







def create_horizontal_bar_chart(labels, percentages, folder_path):
    fig, ax = plt.subplots(figsize=(60, 20))
    coverted_percentages = [i/80*100 for i in percentages][::-1]
    bar_colors = [(COLOR_MAPPING[label.lower()]) for label in labels]
    bars = ax.barh(labels, coverted_percentages, color=bar_colors[::-1], edgecolor='none',height=0.7)
    ax.set_xlim(0, 100)

    for bar, percentage in zip(bars, percentages):
        width = bar.get_width() - 5
        ax.text(width, bar.get_y() + bar.get_height() / 2, f'{percentage}', va='center')

    plt.tight_layout()
    
    plt.savefig(os.path.join(folder_path, "top3dimmensions.png"), dpi=300)
    plt.clf()
    plt.close()



def create_line_chart(virtues, folder_path):
    variables = [i["virtue"] for i in virtues]
    percentage_values = [round(i["rank"]*100/40) for i in virtues]

    x = range(1, len(variables) + 1)
    plt.figure(figsize=(6, 3))
    plt.plot(x, percentage_values, marker='o', linestyle='-', color='b')

    plt.xlabel("Virtues")
    plt.ylabel("Percentage")
    plt.ylim(0,100)
    plt.xticks(x, variables, rotation=45)  # Set variable names as x-axis labels with 45-degree rotation
    plt.grid()


    plt.tight_layout()
    plt.savefig(os.path.join(folder_path, "virtueschart.png"), dpi=300,format="png",  transparent=True)
    plt.clf()
    plt.close()

def insert_image_into_excel(worksheet_name, data=None,excel=None):
    if excel:
        workbook = excel
    else:

        workbook = load_workbook(os.path.join(API_DIR, "assessment","sample", "level2.xlsx"))
    worksheet = workbook[worksheet_name]

    if worksheet_name == 'Page2':
        replacements = {}


        sr_no = 22
        cell_row = ['B48', 'G48', 'L48','B74', 'G74', 'L74']
        for sublist in data:
            for item in sublist:
                # print(item)
                replacements[f"T{sr_no}"] = item['name']
                replacements[f"U{sr_no}"] = round(item['value'])
                if sr_no-22 < 6:
                    replacements[cell_row[sr_no-22]] = ROLES[item['name'].lower()]
                sr_no = sr_no + 1
        
        update_worksheet_cells(worksheet,replacements)
    elif worksheet_name == 'Page3':
        
        replacements = {}

        res_data,nlp_data = data

        cell_row = ['B31','G31','L31']
        for i in range(3):
            replacements[cell_row[i]] = ROLES[res_data[i]['name'].lower()]
        
        cell_row = [('B43','B45','B51','B53'),('B62','B64','B70','B72')]
        for i,j in enumerate(nlp_data):
            replacements[cell_row[i][0]] = j['name'] 
            replacements[cell_row[i][1]] = j['statement'] 
            replacements[cell_row[i][2]] = ' '.join(j['learnings'].split()[:2]) 
            replacements[cell_row[i][3]] = ' '.join(j['learnings'].split()[2:]) 
        
        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name == 'Page4':
        bucket,details,type = data
        bucket_id = bucket['id']
        feature_data = details[bucket_id]

        b70_cell = worksheet['B70'].value.replace('input_emotion',feature_data['emotion']).replace('input_color',feature_data['colour'])

        replacements = {
            'B24': bucket['name'],
            'B30' : feature_data[f'{type}_motivation'],
            'B64' : feature_data['motivation'],
            'B67' : ['input_dimmension',bucket['name']],
            'B69' : ['input_dimmension',bucket['name']],
            'B70' : b70_cell,
            'B72' : ['input_virtue',feature_data['virtue']],
            'B74': feature_data[f'{type}_virtue']
        }

        for i,j in enumerate(feature_data['purpose_statements'].split('\n')):
            replacements[f"B{40+i*2}"] = j

        for i,j in enumerate(feature_data['passion_statements'].split('\n')):
            replacements[f"J{40+i*2}"] = j

        for i,j in enumerate(dict(sorted(bucket['activities'].items(), key=lambda item: item[1], reverse=True)).keys()):
            replacements[f'B{55+i*2}'] = j 

        add_image_to_worksheet(worksheet,graphics_path,f"{bucket['name']}.png",50,12,390,560)
        # add_image_to_worksheet(worksheet,activities_path,f"{bucket['name']}.png",23,6,120,120)
        # add_image_to_worksheet(worksheet,emotions_path,f"{bucket['name']}.png",34,7,120,120)
        # add_image_to_worksheet(worksheet,virtues_path,f"{bucket['name']}.png",38,2,120,120)

        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name in ['Page5','Page6']:
        bucket,details,type = data
        bucket_id = bucket['id']
        feature_data = details[bucket_id]

        b70_cell = worksheet['B70'].value.replace('input_emotion',feature_data['emotion']).replace('input_color',feature_data['colour'])

        replacements = {
            'B24': bucket['name'],
            'B30' : feature_data[f'{type}_motivation'],
            'B64' : feature_data['motivation'],
            'B67' : ['input_dimmension',bucket['name']],
            'B69' : ['input_dimmension',bucket['name']],
            'B70' : b70_cell,
            'B72' : ['input_virtue',feature_data['virtue']],
            'B74': feature_data[f'{type}_virtue']
        }

        for i,j in enumerate(feature_data['purpose_statements'].split('\n')):
            replacements[f"B{40+i*2}"] = j

        for i,j in enumerate(feature_data['passion_statements'].split('\n')):
            replacements[f"J{40+i*2}"] = j

        for i,j in enumerate(dict(sorted(bucket['activities'].items(), key=lambda item: item[1], reverse=True)).keys()):
            replacements[f'B{55+i*2}'] = j 
        
        add_image_to_worksheet(worksheet,graphics_path,f"{bucket['name']}.png",50,12,390,560)
        # add_image_to_worksheet(worksheet,assets_folder,f"{bucket['name']}.png",7,4,250,325)
        # add_image_to_worksheet(worksheet,activities_path,f"{bucket['name']}.png",23,6,120,120)
        # add_image_to_worksheet(worksheet,emotions_path,f"{bucket['name']}.png",34,7,120,120)
        # add_image_to_worksheet(worksheet,virtues_path,f"{bucket['name']}.png",38,2,120,120)

        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name in ['Page7','Page8','Page9','Page10','Page11','Page12']:
        bucket,details,type = data
        bucket_id = bucket['id']
        feature_data = details[bucket_id]


        replacements = {
            'B24': bucket['name'],
            'B30' : feature_data[f'{type}_motivation'],
            'B64' : ['input_virtue',feature_data['virtue']],
            'B66': feature_data[f'{type}_virtue']
        }

        for i,j in enumerate(feature_data['purpose_statements'].split('\n')):
            replacements[f"B{40+i*2}"] = j

        for i,j in enumerate(feature_data['passion_statements'].split('\n')):
            replacements[f"J{40+i*2}"] = j

        for i,j in enumerate(dict(sorted(bucket['activities'].items(), key=lambda item: item[1], reverse=True)).keys()):
            replacements[f'B{55+i*2}'] = j 

        # add_image_to_worksheet(worksheet,assets_folder,f"{bucket['name']}.png",7,4,250,325)
        # add_image_to_worksheet(worksheet,activities_path,f"{bucket['name']}.png",25,6,120,120)
        # add_image_to_worksheet(worksheet,virtues_path,f"{bucket['name']}.png",30,2,120,120)

        add_image_to_worksheet(worksheet,graphics_path,f"{bucket['name']}.png",50,12,390,560)
        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name == 'Page13':

        folder_path,user_profile,user_data = data

        user_top3 = user_data[0]
        user_top3_labels = [i['name'] for i in user_top3]

        generate_user_pie_chart(user_top3_labels,folder_path)
        job_info = generate_jobs_pie_chart(folder_path,user_profile,count=3)

        add_image_to_worksheet(worksheet,folder_path,"job1dimmensions.png",22,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",22,9,500,230)
        
        add_image_to_worksheet(worksheet,folder_path,"job2dimmensions.png",50,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",50,9,500,230)

        replacements = {
            "B20":f"{job_info[0]['job_name']}'s Inclinations",
            "J20":['user_name',user_profile.name],
            "B35":['job_name',job_info[0]['job_name']],
            "J35":['user_name',user_profile.name],
            "B48":f"{job_info[1]['job_name']}'s Inclinations",
            "J48":['user_name',user_profile.name],
            "B63":['job_name',job_info[1]['job_name']],
            "J63":['user_name',user_profile.name],
        }

        for i in range(38, 45,3):
            field_key = f"lwdimension_field{(i - 38) // 3 + 1}"
            dimension_value = job_info[0][field_key]

            replacements[f"B{i}"] = dimension_value
            replacements[f"E{i}"] = ROLES[dimension_value.lower()]

            replacements[f"J{i}"] = dimension_value
            replacements[f"M{i}"] = find_index_of_dimmension(user_data, dimension_value)
        
        
        for i in range(66,73,3):
            field_key = f"lwdimension_field{(i - 66) // 3 + 1}"
            dimension_value = job_info[1][field_key]

            replacements[f"B{i}"] = dimension_value
            replacements[f"E{i}"] = ROLES[dimension_value.lower()]

            replacements[f"J{i}"] = dimension_value
            replacements[f"M{i}"] = find_index_of_dimmension(user_data, dimension_value)
            
        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name == 'Page14':

        folder_path,user_profile,user_data = data

        
        # user_top3_labels = [i['name'] for i in user_top3]

        job_info = generate_jobs_pie_chart(user_profile=user_profile,count=0)

        add_image_to_worksheet(worksheet,folder_path,"job3dimmensions.png",31,9,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",47,9,500,230)

        replacements = {
            "J27":f"{job_info[2]['job_name']}'s Inclinations", 
            "J45":['user_name',user_profile.name],
            "B47":['user_name',user_profile.name],
            "B31":['job_name',job_info[2]['job_name']], 
        }
        for i in range(34, 41,3):
            field_key = f"lwdimension_field{(i - 34) // 3 + 1}"
            dimension_value = job_info[2][field_key]

            replacements[f"B{i}"] = dimension_value
            replacements[f"E{i}"] = ROLES[dimension_value.lower()]

            replacements[f"B{i+16}"] = dimension_value
            replacements[f"E{i+16}"] = find_index_of_dimmension(user_data, dimension_value)
        
 
        update_worksheet_cells(worksheet,replacements)
    return workbook

def convert_excel_to_pdf(folder_path, type, excel_filename="level2report.xlsx", num_pages=14, page_prefix="Page"):
    
    image_paths = []
    # if type == "Career":
    #     pages_to_include = [1, 2, 3,6]
    # elif type == "Leadership":
    #     pages_to_include = [1, 2, 4]
    pages_to_include = list(range(1, num_pages + 1))

    for page_num in pages_to_include:
        page_name = f"{page_prefix}{page_num}"
        image_name = f"{page_name}.png"
        excel2img.export_img(
            os.path.join(folder_path,excel_filename),
            os.path.join(folder_path, image_name),
            page_name,
            "A1:Q77"
        )
        image_paths.append(os.path.join(folder_path, image_name))

   
    pdf_data = img2pdf.convert(image_paths)

    pdf_output_path = os.path.join(folder_path, "output.pdf")
    with open(pdf_output_path, "wb") as file:
        file.write(pdf_data)

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if filename != "output.pdf" and os.path.isfile(file_path):
            os.remove(file_path)