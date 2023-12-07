import os
import uuid
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from api.assessment.constants import ROLES, COLOR_MAPPING
from matplotlib.patheffects import withStroke
import matplotlib.patheffects as path_effects
import matplotlib
matplotlib.use('agg')
from matplotlib import pyplot as plt
import numpy as np
import excel2img
import img2pdf
from multiprocessing import Pool
from api.constants import API_DIR

assets_folder = os.path.join(API_DIR, "assessment","assets","level1")

def add_image_to_worksheet(worksheet, folder_path, image_filename, row, column, width, height):
    img = Image(os.path.join(folder_path, image_filename))
    cell = worksheet.cell(row=row, column=column)
    img.width = width
    img.height = height

    worksheet.add_image(img, cell.coordinate)

def Generate_level1_Report(input_label, input_percentages, inclines,virtues,job_info,user_profile,type='career'):
    
   
    random_hash = str(uuid.uuid4().hex)
    new_folder_path = create_output_folder(random_hash)
    labels = input_label
    percentages = input_percentages

    create_horizontal_bar_chart(labels, percentages, new_folder_path)
    create_line_chart(virtues, new_folder_path)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page3', inclines=inclines[0],user_profile=user_profile)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page4', inclines=inclines[1],excel=excel)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page5', inclines=inclines[2],excel=excel)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page6',excel=excel,virtues=virtues)
    excel = insert_image_into_excel(new_folder_path, labels, percentages, worksheet_name='Page7',excel=excel,job_info=job_info,user_profile=user_profile)
    excel.save(os.path.join(new_folder_path, "level1report.xlsx"))

    convert_excel_to_pdf(new_folder_path,type)

    return os.path.join(new_folder_path, "output.pdf")


def update_worksheet_cells(worksheet, replacements):
    
    for cell_reference, replacement_value in replacements.items():
        
        if isinstance(replacement_value, list):
            param,newValue = replacement_value
            worksheet[cell_reference] = worksheet[cell_reference].value.replace(param,newValue)

        else :
            worksheet[cell_reference] = replacement_value


def create_output_folder(folder_name):
    folder_path = os.path.join(API_DIR, "assessment","reports","level1", folder_name)
    os.makedirs(folder_path)
    return folder_path

def create_horizontal_bar_chart(labels, percentages, folder_path):
    for k in range(1):
        fig, ax = plt.subplots(figsize=(60, 20))
        coverted_percentages = [i/72*100 for i in percentages][::-1]
        bar_colors = [(COLOR_MAPPING[label.lower()]) for label in labels]
        bars = ax.barh(labels, coverted_percentages, color=bar_colors[::-1], edgecolor='none',height=0.7)
        ax.set_xlim(0, 100)

        for i,(bar, percentage) in enumerate(zip(bars, percentages)):
            width = bar.get_width() - 5
            ax.text(width, bar.get_y() + bar.get_height() / 2, f'{percentage}', va='center')

            # if not i ==2-k:
            #     bar.set_alpha(0.1)
            # else:
            bar.set_alpha(1)

        plt.tight_layout()
        
        plt.savefig(os.path.join(folder_path, f"top3dimmensions-{k+1}.png"))
        plt.clf()
        plt.close()



def create_line_chart(virtues, folder_path):
    variables = [i["virtue"] for i in virtues]
    percentage_values = [round(i["rank"]*100/40) for i in virtues]

    x = range(1, len(variables) + 1)
    plt.figure(figsize=(6, 3))
    plt.plot(x, percentage_values, marker='o', linestyle='-', color='b')

    plt.xlabel("Virtues")
    plt.ylabel("Percentage")
    plt.ylim(0,100)
    plt.xticks(x, variables, rotation=45)  # Set variable names as x-axis labels with 45-degree rotation4

    y_ticks = [i for i in range(0, 101, 10)]

    plt.grid()


    plt.tight_layout()
    plt.savefig(os.path.join(folder_path, "virtueschart.png"), dpi=300,format="png",  transparent=True)
    plt.clf()
    plt.close()

def insert_image_into_excel(folder_path, labels, percentages, worksheet_name, inclines=None,excel=None,virtues=None,job_info=[],user_profile=None):
    if excel:
        workbook = excel
    else:

        workbook = load_workbook(os.path.join(API_DIR, "assessment","sample", "level1.xlsx"))
    worksheet = workbook[worksheet_name]

    

    if worksheet_name == 'Page3':
        add_image_to_worksheet(worksheet,folder_path,"top3dimmensions-1.png",12,12,330,140)
        replacements = {
            "C15":user_profile.name,
            "C19":user_profile.dob,
            "C23":user_profile.gender,
            "C27":user_profile.marital_status,
            "C31":user_profile.primary_mobile_no,
            "C35":user_profile.user.email,
            "C40":user_profile.residential_address,
            "C60":user_profile.goals,
            "H12": labels[0],
            "H14": labels[1],
            "H17": labels[2],
            "K12": round(percentages[0] * 100 / 72),
            "K14": round(percentages[1] * 100 / 72),
            "K17": round(percentages[2] * 100 / 72)
        }
        
        for index,job in enumerate(user_profile.job_aspirations.all()):
            replacements[f"C{47+index*2}"] = job.title
        
        update_worksheet_cells(worksheet,replacements)
        update_page2_cells(worksheet, inclines)
    
    
    
    elif worksheet_name in ['Page4','Page5']:
        # add_image_to_worksheet(worksheet,folder_path,"top3dimmensions-2.png",7,7,330,130)

        replacements = {
            # "D7": labels[0],
            # "D9": labels[1],
            # "D11": labels[2],
            # "F7": round(percentages[0] * 100 / 72),
            # "F9": round(percentages[1] * 100 / 72),
            # "F11": round(percentages[2] * 100 / 72)
        }

        update_worksheet_cells(worksheet,replacements)
        update_page3_cells(worksheet, inclines)

    elif worksheet_name == 'Page6':
        add_image_to_worksheet(worksheet,folder_path,"virtueschart.png",14,10,400,200)
        update_page4_cells(worksheet,virtues,folder_path)

    
    elif worksheet_name == 'Page7':
        update_page5_cells(worksheet,folder_path,labels,job_info,user_profile)
        

    
    return workbook

    

def update_page2_cells(worksheet, inclines):
    replacements = {
            "G25": inclines['feature'],
            "H34":inclines['purpose_statement'],
            "H42":"You " + inclines['thrive_environment'].lower(),
            "H64":inclines['career_inclination_statement'],
            "I71":inclines['quote'],
    }
    
    

    inclinations = inclines['inclinations'].split('\n')
   
    for i in range(len(inclinations)):
        replacements[f"H{50+i*2}"] = inclinations[i] 
    
    
    update_worksheet_cells(worksheet,replacements)
    add_image_to_worksheet(worksheet,assets_folder,f"{inclines['feature']}.png",17,13,300,400)

def update_page3_cells(worksheet, inclines):
    # pass
    replacements = {
            "A20": inclines['feature'],
            "B31":inclines['purpose_statement'],
            "B39":"You " + inclines['thrive_environment'].lower(),
            "B65":inclines['career_inclination_statement'],
            "J64":inclines['quote']
    }



    inclinations = inclines['inclinations'].split('\n')
    for i in range(len(inclinations)):
        replacements[f"C{47+i*2}"] = inclinations[i]
    
    

    update_worksheet_cells(worksheet,replacements)
    add_image_to_worksheet(worksheet,assets_folder,f"{inclines['feature']}.png",8,7,400,550)
    


def update_page6_cells(worksheet, inclines):

    replacements = {
        "B7":["input_dimension",inclines['feature']],
        "B13":inclines['purpose_statement'],
        "B16":"You "+ inclines['thrive_environment'],
        "B34":inclines['career_inclination_statement'],
        "B40":inclines['quote']


    }
    inclinations = inclines['inclinations'].split('\n')
    cont = 21
    for i in range(len(inclinations)):
        replacements[f"B{cont}"] = inclinations[i]
        cont = cont + 2 

    # careers = inclines['careers'].split(",")[:5]
    # for i in range(37, 40):
    #     replacements[f"C{i}"] = careers[i - 37]
    
    update_worksheet_cells(worksheet,replacements)


def update_page4_cells(worksheet,virtues,folder_path):
    cell_positions = [("B25", "B23"), ("B35", "B33"), ("B45", "B43")]
    

    for i, (value_cell, virtue_cell) in enumerate(cell_positions):
        worksheet[value_cell] = virtues[i]['text']
        worksheet[virtue_cell] = virtues[i]['virtue']
        worksheet[f"S{i+11}"] = round(virtues[i]["rank"]*100/40)


    
    middle_three = []
    cells = ('57','60','63')
    for i in range(3):
        worksheet[f"C{cells[i]}"] = f'{virtues[i+3]["virtue"]}'
        middle_three.append(virtues[i+3]["virtue"])
    
    bottom_three = []
    cells = ('68','71','74')
    for i in range(3):
        worksheet[f"C{cells[i]}"] = f'{virtues[i+6]["virtue"]}'
        bottom_three.append(virtues[i+6]["virtue"])

    replacements = {
             "J57": ['middle_three_virtues',', '.join(middle_three[:-1]) + ' and ' + middle_three[-1]],
             "J68": ['bottom_three_virtues',', '.join(bottom_three[:-1]) + ' and ' + bottom_three[-1]],

    }
    
    update_worksheet_cells(worksheet,replacements)


def update_page5_cells(worksheet,folder_path,labels,job_info,user_profile):


    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    recipe = list(COLOR_MAPPING.keys())
    data = [1]*len(recipe)

    colors = [COLOR_MAPPING[recipe_name] for recipe_name in recipe]

    # Create an "explode" list to specify which segments to explode
    explode = [0.05, 0.05, 0.05, 0.05, 0.05, 0.05,0.05, 0.05, 0.05]

    wedges, texts = ax.pie(data, colors=colors, wedgeprops=dict(width=0.15), startangle=-40, explode=explode)

    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-"), bbox=bbox_props, zorder=0, va="center")

    # Add annotations only for "angel" and "principal"

    lw_dimensions = job_info[0]
    dimension_fields = ['lwdimension_field1', 'lwdimension_field2', 'lwdimension_field3']

    annotation_indices = [recipe.index(lw_dimensions[dim].lower()) for dim in dimension_fields]

    

    for i, p in enumerate(wedges):
        if i in annotation_indices:
            ang = (p.theta2 - p.theta1) / 2. + p.theta1
            y = np.sin(np.deg2rad(ang))
            x = np.cos(np.deg2rad(ang))
            horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
            connectionstyle = f"angle,angleA=0,angleB={ang}"
            kw["arrowprops"].update({"connectionstyle": connectionstyle})
            ax.annotate(recipe[i], xy=(x, y), xytext=(1.75 * np.sign(x), 1.4 * y),
                        horizontalalignment=horizontalalignment, **kw)
            
            worksheet[f"B{48+annotation_indices.index(i)*3}"] = recipe[i].capitalize()
            worksheet[f"E{48+annotation_indices.index(i)*3}"] = ROLES[recipe[i].lower()]

    center_text = plt.gca().text(0.0, 0.0, 'Limeneal', color='white', ha='center', va='center', size=10,fontfamily='sans-serif')
    stroke = withStroke(linewidth=3, foreground='black')
    center_text.set_path_effects([stroke, path_effects.Normal()])

    center_circle = plt.Circle((0, 0), 0.35, color='#f5cff7')
    ax.add_artist(center_circle)
    plt.tight_layout()

    plt.savefig(os.path.join(folder_path, "job1dimmensions.png"), dpi=300,transparent=True)
    plt.clf()
    plt.close()
    
    add_image_to_worksheet(worksheet,folder_path,"job1dimmensions.png",29,2,500,230)

    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))
    explode = [0.05, 0.05, 0.05, 0.05, 0.05, 0.05,0.05, 0.05, 0.05]

    wedges, texts = ax.pie(data, colors=colors, wedgeprops=dict(width=0.15), startangle=-40, explode=explode)

    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-"), bbox=bbox_props, zorder=0, va="center")


    lw_dimensions = job_info[0]
    

    annotation_indices = [recipe.index(i.lower()) for i in labels]

    for i, p in enumerate(wedges):
        if i in annotation_indices:
            ang = (p.theta2 - p.theta1) / 2. + p.theta1
            y = np.sin(np.deg2rad(ang))
            x = np.cos(np.deg2rad(ang))
            horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
            connectionstyle = f"angle,angleA=0,angleB={ang}"
            kw["arrowprops"].update({"connectionstyle": connectionstyle})
            ax.annotate(recipe[i], xy=(x, y), xytext=(1.75 * np.sign(x), 1.4 * y),
                        horizontalalignment=horizontalalignment, **kw)

           
    center_text = plt.gca().text(0.0, 0.0, 'Limeneal', color='white', ha='center', va='center', size=10,fontfamily='sans-serif')
    stroke = withStroke(linewidth=3, foreground='black')
    center_text.set_path_effects([stroke, path_effects.Normal()])

    center_circle = plt.Circle((0, 0), 0.35, color='#f5cff7')
    ax.add_artist(center_circle)
    # plt.title(f"Sumith's Inclination'")
    plt.tight_layout()

    plt.savefig(os.path.join(folder_path, "usersdimmensions.png"), dpi=300,transparent=True)
    plt.clf()
    plt.close()

    add_image_to_worksheet(worksheet,folder_path,"usersdimmensions.png",29,10,500,230)
    
    replacements = {
        "B22": f"{job_info[0]['job_name']}'s Inclinations",
        "J22": f"{user_profile.name}'s Inclinations",
        "B45": ['job_name',job_info[0]['job_name']],
        "J45": ['user_name',user_profile.name]
    }

    update_worksheet_cells(worksheet,replacements)
    for i in range(3):
        worksheet[f"J{48+i*3}"] = labels[i].capitalize()
        worksheet[f"M{48+i*3}"] =   ROLES[labels[i].lower()]

def convert_page_to_image(args):
    folder_path, excel_filename, page_num, page_prefix = args
    page_name = f"{page_prefix}{page_num}"
    image_name = f"{page_name}.png"

    excel2img.export_img(
        os.path.join(folder_path, excel_filename),
        os.path.join(folder_path, image_name),
        "",
        f"{page_name}!A1:Q77"
    )

    return os.path.join(folder_path, image_name)

def convert_excel_to_pdf(folder_path, type, excel_filename="level1report.xlsx", num_pages=8, page_prefix="Page"):
    # Determine pages to include based on type or use all pages
    # if type == "Career":
    #     pages_to_include = [1, 2, 3, 6, 4, 5]
    # elif type == "Leadership":
    #     pages_to_include = [1, 2, 4]
    # else:
    pages_to_include = list(range(1, num_pages + 1))

    pool = Pool()

    args_list = [(folder_path, excel_filename, page_num, page_prefix) for page_num in pages_to_include]

    image_paths = pool.map(convert_page_to_image, args_list)

    # Close the pool to release resources
    pool.close()
    pool.join()

    pdf_data = img2pdf.convert(image_paths)

    pdf_output_path = os.path.join(folder_path, "output.pdf")
    with open(pdf_output_path, "wb") as file:
        file.write(pdf_data)

    for filename in image_paths:
        os.remove(filename)