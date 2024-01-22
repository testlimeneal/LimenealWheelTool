import os
import uuid
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from api.assessment.constants import ROLES, COLOR_MAPPING, LEVEL3_ROLES

from api.assessment.utils.common import generate_user_pie_chart,generate_jobs_pie_chart
import matplotlib
matplotlib.use('agg')
from matplotlib import pyplot as plt
from matplotlib.patheffects import withStroke
import matplotlib.patheffects as path_effects
import numpy as np
import excel2img
import img2pdf
from api.constants import API_DIR
from api.assessment.helperfunctions.common import get_careers_from_dimmensions, generate_report_file_path


activities_path = os.path.join(API_DIR, "assessment","assets","level2", 'dimmensions')
assets_folder = os.path.join(API_DIR, "assessment","assets","level1")

emotions_path = os.path.join(API_DIR, "assessment","assets","level2", 'emotions')
virtues_path = os.path.join(API_DIR, "assessment","assets","level2", 'virtues')


def find_index_of_dimmension(data, target_name):
    role_index = data.index(target_name)

        # Determine preference level
    if role_index in range(0, 3):
        preference_level = "Most Preferred"
    elif role_index in range(3, 6):
        preference_level = "Lesser Preferred"
    elif role_index in range(6, 9):
        preference_level = "Least Preferred"
    
    return preference_level


def add_image_to_worksheet(worksheet, folder_path, image_filename, row, column, width, height):
    img = Image(os.path.join(folder_path, image_filename))
    cell = worksheet.cell(row=row, column=column)
    img.width = width
    img.height = height
    worksheet.add_image(img, cell.coordinate)

def create_output_folder(folder_name):
    folder_path = os.path.join(API_DIR, "assessment","reports","level3", folder_name)
    os.makedirs(folder_path)
    return folder_path

def update_worksheet_cells(worksheet, replacements):
    for cell_reference, replacement_value in replacements.items():
        
        if isinstance(replacement_value, list):
            param,newValue = replacement_value
            worksheet[cell_reference] = worksheet[cell_reference].value.replace(param,newValue)

        else :
            worksheet[cell_reference] = replacement_value



def Generate_level3_Report(user_profile,data):
    
    random_hash = str(uuid.uuid4().hex)
    new_folder_path = create_output_folder(random_hash)
    
   

    user_dimmensions = [i[0] for i in data]
    generate_user_pie_chart(user_dimmensions[0:3],new_folder_path)
    job_info = generate_jobs_pie_chart(new_folder_path,user_profile,count=5)
    
    excel = insert_image_into_excel(worksheet_name='Data',data=data)
    
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page2',data=user_dimmensions[0:2])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page3',data=user_dimmensions[2:3])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page4',data=user_dimmensions[3:5])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page5',data=user_dimmensions[5:6])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page6',data=user_dimmensions[6:8])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page7',data=user_dimmensions[8:9])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page8',data=(user_profile,job_info,user_dimmensions),folder_path=new_folder_path)
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page9',data=(user_profile,job_info,user_dimmensions),folder_path=new_folder_path)
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page10',data=(user_profile,job_info,user_dimmensions),folder_path=new_folder_path)
    
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page11',data=user_dimmensions[0:6])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page12',data=user_dimmensions[6:9])
    excel = insert_image_into_excel(excel=excel,worksheet_name='Page13',data=data)

    excel.save(os.path.join(new_folder_path, "level3report.xlsx"))

    convert_excel_to_pdf(new_folder_path,"type")

    return os.path.join(new_folder_path, "output.pdf")







def create_horizontal_bar_chart(labels, percentages, folder_path):
    fig, ax = plt.subplots(figsize=(60, 20))
    coverted_percentages = [i/80*100 for i in percentages][::-1]
    bar_colors = [(COLOR_MAPPING[label.lower()]) for label in labels]
    bars = ax.barh(labels, coverted_percentages, color=bar_colors[::-1], edgecolor='none',height=0.7)
    ax.set_xlim(0, 100)

    for bar, percentage in zip(bars, percentages):
        width = bar.get_width() - 5
        ax.text(width, bar.get_y() + bar.get_height() / 2, f'{percentage}', va='center')

    plt.tight_layout()
    
    plt.savefig(os.path.join(folder_path, "top3dimmensions.png"), dpi=300)
    plt.clf()
    plt.close()



def create_line_chart(virtues, folder_path):
    variables = [i["virtue"] for i in virtues]
    percentage_values = [round(i["rank"]*100/40) for i in virtues]

    x = range(1, len(variables) + 1)
    plt.figure(figsize=(6, 3))
    plt.plot(x, percentage_values, marker='o', linestyle='-', color='b')

    plt.xlabel("Virtues")
    plt.ylabel("Percentage")
    plt.ylim(0,100)
    plt.xticks(x, variables, rotation=45)  # Set variable names as x-axis labels with 45-degree rotation
    plt.grid()


    plt.tight_layout()
    plt.savefig(os.path.join(folder_path, "virtueschart.png"), dpi=300,format="png",  transparent=True)
    plt.clf()
    plt.close()

def insert_image_into_excel(worksheet_name, data=None,excel=None,folder_path=None):
    if excel:
        workbook = excel
    else:

        workbook = load_workbook(os.path.join(API_DIR, "assessment","sample", "level3.xlsx"))
    worksheet = workbook[worksheet_name]

    # return workbook
    if worksheet_name == 'Data':
        replacements = {}
        row = 2
        for i,dimmension in enumerate(data):
            replacements[f"A{i+2}"] = dimmension[0]
            replacements[f"B{i+2}"] = ROLES[dimmension[0].lower()]
            replacements[f"C{i+2}"] = dimmension[1] 
           
            replacements[f"E{row}"] = dimmension[2][0]
            replacements[f"F{row}"] = dimmension[2][1]
            replacements[f"G{row}"] = dimmension[2][2]


            row = row + 1
            for j in dimmension[3]:
                replacements[f"E{row}"] = j[0]
                replacements[f"F{row}"] = j[1]
                replacements[f"G{row}"] = j[2]
                row = row + 1
            

        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name in ['Page2','Page4','Page6']:
     
        add_image_to_worksheet(worksheet,virtues_path,f"{data[0]}.png",20,2,90,90) 
        add_image_to_worksheet(worksheet,virtues_path,f"{data[1]}.png",50,2,90,90) 

    elif worksheet_name == ['Page3','Page5','Page7']:
        add_image_to_worksheet(worksheet,virtues_path,f"{data[0]}.png",20,2,90,90) 
       
    elif worksheet_name == 'Page8': 
        user_profile,job_info,user_data = data  

        add_image_to_worksheet(worksheet,folder_path,"job1dimmensions.png",22,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",22,10,500,230)

        add_image_to_worksheet(worksheet,folder_path,"job2dimmensions.png",51,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",51,10,500,230)

        replacements = {
            "B20":f"{job_info[0]['job_name']}'s Inclinations",
            "J20":['user_name',user_profile.name],
            "B35":['job_name',job_info[0]['job_name']],
            "J35":['user_name',user_profile.name],

            "B48":f"{job_info[1]['job_name']}'s Inclinations",
            "J48":['user_name',user_profile.name],
            "B63":['job_name',job_info[1]['job_name']],
            "J63":['user_name',user_profile.name],
           
        }

        for i,j in enumerate(range(38,45,3)):
            field_key = f"lwdimension_field{i + 1}"
            dimension_value = job_info[0][field_key]

            replacements[f"B{j}"] = dimension_value
            replacements[f"E{j}"] = ROLES[dimension_value.lower()]

            replacements[f"J{j}"] = dimension_value
            replacements[f"M{j}"] = find_index_of_dimmension(user_data, dimension_value)
        
        for i,j in enumerate(range(66,73,3)):
            field_key = f"lwdimension_field{i + 1}"
            dimension_value = job_info[1][field_key]

            replacements[f"B{j}"] = dimension_value
            replacements[f"E{j}"] = ROLES[dimension_value.lower()]

            replacements[f"J{j}"] = dimension_value
            replacements[f"M{j}"] = find_index_of_dimmension(user_data, dimension_value)
        
        

        update_worksheet_cells(worksheet,replacements)

    elif worksheet_name == 'Page9': 
        user_profile,job_info,user_data = data  

        add_image_to_worksheet(worksheet,folder_path,"job3dimmensions.png",22,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",22,10,500,230)

        add_image_to_worksheet(worksheet,folder_path,"job4dimmensions.png",51,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",51,10,500,230)

        replacements = {
            "B20":f"{job_info[2]['job_name']}'s Inclinations",
            "J20":['user_name',user_profile.name],
            "B35":['job_name',job_info[2]['job_name']],
            "J35":['user_name',user_profile.name],

            "B48":f"{job_info[3]['job_name']}'s Inclinations",
            "J48":['user_name',user_profile.name],
            "B63":['job_name',job_info[3]['job_name']],
            "J63":['user_name',user_profile.name],
           
        }

        for i,j in enumerate(range(38,45,3)):
            field_key = f"lwdimension_field{i + 1}"
            dimension_value = job_info[2][field_key]

            replacements[f"B{j}"] = dimension_value
            replacements[f"E{j}"] = ROLES[dimension_value.lower()]

            replacements[f"J{j}"] = dimension_value
            replacements[f"M{j}"] = find_index_of_dimmension(user_data, dimension_value)
        
        for i,j in enumerate(range(66,73,3)):
            field_key = f"lwdimension_field{i + 1}"
            dimension_value = job_info[3][field_key]

            replacements[f"B{j}"] = dimension_value
            replacements[f"E{j}"] = ROLES[dimension_value.lower()]

            replacements[f"J{j}"] = dimension_value
            replacements[f"M{j}"] = find_index_of_dimmension(user_data, dimension_value)
        
        

        update_worksheet_cells(worksheet,replacements)
    
    elif worksheet_name == 'Page10': 
        user_profile,job_info,user_data = data  

        add_image_to_worksheet(worksheet,folder_path,"job5dimmensions.png",22,2,500,230)
        add_image_to_worksheet(worksheet,folder_path,"user_top3.png",22,10,500,230)

        replacements = {
            "B20":f"{job_info[4]['job_name']}'s Inclinations",
            "J20":['user_name',user_profile.name],
            "B35":['job_name',job_info[4]['job_name']],
            "J35":['user_name',user_profile.name],

           
        }

        for i,j in enumerate(range(38,45,3)):
            field_key = f"lwdimension_field{i + 1}"
            dimension_value = job_info[4][field_key]

            replacements[f"B{j}"] = dimension_value
            replacements[f"E{j}"] = ROLES[dimension_value.lower()]

            replacements[f"J{j}"] = dimension_value
            replacements[f"M{j}"] = find_index_of_dimmension(user_data, dimension_value)
        update_worksheet_cells(worksheet,replacements)
        
    
    
    
    elif worksheet_name == 'Page11': 
        add_image_to_worksheet(worksheet,assets_folder,f"{data[0].lower()}.png",14,1,330,450)
        add_image_to_worksheet(worksheet,assets_folder,f"{data[1].lower()}.png",14,7,330,450)
        add_image_to_worksheet(worksheet,assets_folder,f"{data[2].lower()}.png",14,12,330,450)

        add_image_to_worksheet(worksheet,assets_folder,f"{data[3].lower()}.png",44,1,330,450)
        add_image_to_worksheet(worksheet,assets_folder,f"{data[4].lower()}.png",44,7,330,450)
        add_image_to_worksheet(worksheet,assets_folder,f"{data[5].lower()}.png",44,12,330,450)
        
        replacements = {}
        
        power_careers = get_careers_from_dimmensions(data[0:3])
        push_careers = get_careers_from_dimmensions(data[3:6])


        cells=('B38', 'B41', 'H38', 'H41', 'M38')
        for count,job in enumerate(power_careers):
            replacements[cells[count]] = job.title
            print(cells[count],job.title)
        
        
        cells=('B68', 'B71', 'H68', 'H71', 'M68')
        for count,job in enumerate(push_careers):
            replacements[cells[count]] = job.title

        top_three_labels = [i for i in data[0:3]]
        # top_three_roles = [LEVEL3_ROLES[i.lower()] for i in top_three_labels]

        middle_three_labels = [i for i in data[3:6]]
        cells = ('B','G','L')
        for i in range(3):
            replacements[f"{cells[i]}31"] = top_three_labels[i]
            replacements[f"{cells[i]}61"] = middle_three_labels [i]

        # middle_three_roles = [LEVEL3_ROLES[i.lower()] for i in middle_three_labels]

        # power_sentence = f"Strong in {', '.join(top_three_labels[:-1])}, and {top_three_labels[-1]}, careers where you will find complete fulfilment will be where the inclination is to  {', '.join(top_three_roles[:-1])}, and {top_three_roles[-1]}"
        # push_sentence = f"Strong in {', '.join(middle_three_labels[:-1])}, and {middle_three_labels[-1]}, careers where you will find moderate fulfilment will be where the inclination is to  {', '.join(middle_three_roles[:-1])}, and {middle_three_roles[-1]}"
      
        # replacements['B33'] = power_sentence
        # replacements['B65'] = push_sentence
        update_worksheet_cells(worksheet,replacements)
       
        # for i in enumerate

    elif worksheet_name == 'Page12': 
        add_image_to_worksheet(worksheet,assets_folder,f"{data[0].lower()}.png",14,1,330,450)
        add_image_to_worksheet(worksheet,assets_folder,f"{data[1].lower()}.png",14,7,330,450)
        add_image_to_worksheet(worksheet,assets_folder,f"{data[2].lower()}.png",14,12,330,450)

        replacements = {}

        pain_careers = get_careers_from_dimmensions(data[0:3])
                        
        cells=('B38', 'B41', 'H38', 'H41', 'M38')
        for count,job in enumerate(pain_careers):
            replacements[cells[count]] = job.title
        
        bottom_three_labels = [i for i in data[0:3]]
        cells = ('B','G','L')
        for i in range(3):
            replacements[f"{cells[i]}31"] = bottom_three_labels[i]
        # bottom_three_roles = [LEVEL3_ROLES[i.lower()] for i in bottom_three_labels]

        # pain_sentence = f"Strong in {', '.join(bottom_three_labels[:-1])}, and {bottom_three_labels[-1]}, careers where you will find least fulfilment will be where the inclination is to  {', '.join(bottom_three_roles[:-1])}, and {bottom_three_roles[-1]}"

        # replacements['B33'] = pain_sentence
        
        update_worksheet_cells(worksheet,replacements)
        

    elif worksheet_name == 'Page13':
        
        worksheet = workbook['Page12'] #Did this because code was already implemeneted
        replacements = {}
        rows = {
            "Binder":"V4",
            "Charmer":"V5",
            "Dominion":"V6",
            "Angel":"V7",
            "Mentor":"V8",
            "Principal":"V9",
            "Guardian":"V10",
            "Harmonizer":"V11",
            "Visualizer":"V12",
        }

        values = {}

        for feature in data:
            replacements[rows[feature[0]]] = feature[1]
            values[feature[0]] = feature[1]
        
        # print(values)
        def calculate_average(rating1, rating2):
            return (values[rating1] + values[rating2]) / 2
        averages = {
            'S': calculate_average('Visualizer', 'Mentor'),
            'A': calculate_average('Dominion', 'Charmer'),
            'L': calculate_average('Guardian', 'Principal'),
            'E': calculate_average('Angel', 'Binder'),
            'M': calculate_average('Harmonizer', 'Mentor')
        }
        salem_cells = {
            'S':'E51',
            'A':'E55',
            'L':'E58',
            'E':'E61',
            'M':'E64',
        }
        sorted_data = sorted(averages.items(), key=lambda x: x[1], reverse=True)
        for i, (letter, value) in enumerate(sorted_data):
            if i < 2:
                replacements[salem_cells[letter]] = 'Power'
            elif i < 4:
                replacements[salem_cells[letter]] = 'Push'
            else:
                replacements[salem_cells[letter]] = 'Pain'

        update_worksheet_cells(worksheet,replacements)
        
    return workbook

def convert_excel_to_pdf(folder_path, type, excel_filename="level3report.xlsx", num_pages=15, page_prefix="Page"):
    
    image_paths = []
    # if type == "Career":
    #     pages_to_include = [1, 2, 3,6]
    # elif type == "Leadership":
    #     pages_to_include = [1, 2, 4]
    pages_to_include = list(range(1, num_pages + 1))

    for page_num in pages_to_include:
        page_name = f"{page_prefix}{page_num}"
        image_name = f"{page_name}.png"
        excel2img.export_img(
            os.path.join(folder_path,"level3report.xlsx"),
            os.path.join(folder_path, image_name),
            page_name,
            "A1:Q77"
        )
        image_paths.append(os.path.join(folder_path, image_name))   

   
    pdf_data = img2pdf.convert(image_paths)

    pdf_output_path = os.path.join(folder_path, "output.pdf")
    with open(pdf_output_path, "wb") as file:
        file.write(pdf_data)

    # for filename in os.listdir(folder_path):
    #     file_path = os.path.join(folder_path, filename)
    #     if filename != "output.pdf" and os.path.isfile(file_path):
    #         os.remove(file_path)