import os
from openpyxl import load_workbook
from openpyxl import Workbook
from api.assessment.models import UserResponse,Question,Answer,UserProfile,Level2Response,Level2Question
from itertools import groupby
from api.assessment.helperfunctions.level1 import process_level1_scores
from api.assessment.helperfunctions.level2 import process_level2_scores
from api.assessment.helperfunctions.common import get_feature_name_by_id
import secrets

def generate_excel_report(user_id):
    

    filename = "LimenealScores.xlsx"
    script_directory = os.path.dirname(os.path.abspath(__file__))  

    file_path = os.path.join(script_directory, filename)
    original_workbook = load_workbook(file_path)
    temp_dir = 'temp_folder'
    os.makedirs(temp_dir, exist_ok=True)

    random_hex_code = secrets.token_hex(4)
    temp_file_path = os.path.join(temp_dir, f'temp_report_{random_hex_code}_{os.path.basename(file_path)}')
    original_workbook.save(temp_file_path)

    temp_workbook = load_workbook(temp_file_path)

    level1_scores = UserResponse.objects.filter(user_id=user_id).values()
    sorted_data = sorted(level1_scores, key=lambda x: x['question_id'])

    grouped_data = {
        key: [{'answer_id': item['answer_id'], 'rank': item['rank']} for item in group]
        for key, group in groupby(sorted_data, key=lambda x: x['question_id'])
    }
    temp = []

    cell_data = [
    {'start_cell': 'A1', 'range': ['A', 'D', 5]},
    {'start_cell': 'G1', 'range': ['G', 'J', 5]},
    {'start_cell': 'M1', 'range': ['M', 'P', 5]},
    {'start_cell': 'S1', 'range': ['S', 'V', 5]},
    {'start_cell': 'A15', 'range': ['A', 'D', 19]},
    {'start_cell': 'G15', 'range': ['G', 'J', 19]},
    {'start_cell': 'M15', 'range': ['M', 'P', 19]},
    {'start_cell': 'S15', 'range': ['S', 'V', 19]},
]

    worksheet = temp_workbook['Level1']
    for curr_question,(question_id, answers) in enumerate(grouped_data.items()):
        # Fetch and display the question text
        worksheet[cell_data[curr_question]['start_cell']] = Question.objects.get(id=question_id).text

        # Display each answer text
        for curr_answer,answer in enumerate(answers):
            answer_text = Answer.objects.get(id=answer['answer_id']).text
            worksheet[f"{cell_data[curr_question]['range'][0]}{str(cell_data[curr_question]['range'][2]+curr_answer)}"] = answer_text
            worksheet[f"{cell_data[curr_question]['range'][1]}{str(cell_data[curr_question]['range'][2]+curr_answer)}"] = answer['rank']
    
    user_profile = UserProfile.objects.get(user=user_id)
    if not user_profile.level1:
        user_profile = process_level1_scores(user_id)

    filtered_dict = {key: value for key, value in user_profile.level1.items() if key not in ['bucket', 'virtue']}

    for idx,(id,value) in enumerate(filtered_dict.items()):
        worksheet[f"A{32+idx}"] = get_feature_name_by_id(id)
        worksheet[f"D{32+idx}"] = user_profile.level1['bucket'][id]
        worksheet[f"F{32+idx}"] = value
    


    if not user_profile.level2:
        user_profile = process_level2_scores(user_id)
    
    worksheet = temp_workbook['Level2']
    responses = Level2Response.objects.filter(user=user_id)

    dimmensions = [response for response in responses if response.nlp is None]
    nlp = [response for response in responses if response.nlp is not None]

    grouped_responses = {key: list(group) for key, group in groupby(dimmensions, key=lambda x: x.question)}
    
    cell_data = [
    {'start_cell': 'A1', 'range': ['A', 'D','G', 4]},
    {'start_cell': 'K1', 'range': ['K', 'N','Q', 4]},
    {'start_cell': 'A14', 'range': ['A', 'D','G', 17]},
]
    

    for idx, (question, answer) in enumerate(grouped_responses.items()):
        worksheet[cell_data[idx]['start_cell']] = question.text

        sorted_answer = sorted(answer, key=lambda x: x.rank, reverse=True)
        reshaped_matrix = [sorted_answer[i:i+3] for i in range(0, len(sorted_answer), 3)]

        for inner_idx, response in enumerate(reshaped_matrix):
            for j in range(3):
                col_idx = cell_data[idx]['range'][j]
                worksheet[f"{col_idx}{inner_idx + cell_data[idx]['range'][3]}"] = f"{response[j].rank} - {response[j].answer.text}"
    

    for idx,i in enumerate(nlp):
        level2_question = Level2Question.objects.get(id=i.question.id)
        visual_option = getattr(level2_question,f'{i.nlp}_option')
        worksheet[f"K{15+idx}"] = i.question.text
        worksheet[f"P{15+idx}"] = f"{visual_option} | {i.nlp}" 
    

    idx = 0
    for i in range(3):
        for j in range(3):
            
            worksheet[f"A{31+idx}"] = user_profile.level2['value'][i][j]['name']
            worksheet[f"D{31+idx}"] = user_profile.level2['value'][i][j]['value']
            idx = idx + 1

    score_file = os.path.join(temp_dir, f'report_{random_hex_code}_{os.path.basename(file_path)}')
    temp_workbook.save(score_file)

    os.remove(temp_file_path)
    # # Create a temporary directory if not exists
    # temp_dir = 'temp_reports'
    # os.makedirs(temp_dir, exist_ok=True)

    # # Create a temporary copy of the original workbook
    # temp_file_path = os.path.join(temp_dir, f'temp_report_{os.path.basename(original_file_path)}')
    # original_workbook.save(temp_file_path)

    # # Access the active sheet of the temporary workbook (assuming there is only one sheet, modify accordingly)
    # temp_workbook = load_workbook(temp_file_path)
    # temp_sheet = temp_workbook.active

    # # Add/Edit data in the sheet (replace with your actual data modifications)
    # # For example, adding a new row with data
    # temp_sheet.append(['New Name', 28, 88])

    # # Save the changes to the temporary workbook
    # temp_workbook.save(temp_file_path)

    # # Close the temporary workbook
    # temp_workbook.close()

    # # Return the path to the temporary file
    return score_file